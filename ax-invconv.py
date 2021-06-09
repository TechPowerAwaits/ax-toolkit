# Copyright 2021 Richard Johnston <techpowerawaits@outlook.com>
# SPDX-license-identifier: 0BSD

import argparse
import csv
from openpyxl import load_workbook
import sys

with open("VERSION", "r") as version_file:
    ver_str = version_file.readline()

AXELOR_CSV_COLUMNS = [
    "procurementMethodSelect",
    "purchasesUnit_importId",
    "importId",
    "productSubTypeSelect",
    "purchaseCurrency_importId",
    "productTypeSelect",
    "salePrice",
    "picture_importId",
    "managPriceCoef",
    "purchasePrice",
    "defaultSupplierPartner_importId",
    "internalDescription",
    "salesUnit_importId",
    "name",
    "productFamily_importId",
    "code",
    "description",
    "productCategory_importId",
    "saleSupplySelect",
    "fullName",
    "saleCurrency_importId"
]
AXELOR_FAMILY_SHORTHAND = {
    "Components" : "COMP",
    "Consumables" : "CONS",
    "Equipment" : "EQPT",
    "Expenses" : "TEXP",
    "Services" : "SERV",
    "Subscription" : "HSGT"
}
AXELOR_PRODUCT_CATEGORIES = {
    "Accomodation" : 16,
    "Case" : 7,
    "Hard Disk" : 5,
    "Hosting" : 1,
    "Maintenance" : 2,
    "Meal" : 15,
    "Memory" : 8,
    "Motherboard" : 9,
    "Non-perishable" : 10,
    "Package" : 18,
    "Perishable" : 11,
    "Printer" : 4,
    "Processor" : 6,
    "Project based": 13,
    "Screen" : 17,
    "Server" : 3,
    "Time based" : 12,
    "Transport" : 14
}
AXELOR_CATEGORY_SHORTHAND = {
    "Accomodation" : "ACCO",
    "Case" : "CASE",
    "Hard Disk" : "HDD",
    "Hosting" : "HOST",
    "Maintenance" : "MAIN",
    "Meal" : "MEAL",
    "Memory" : "MEM",
    "Motherboard" : "MOB",
    "Non-perishable" : "NPER",
    "Package" : "PACK",
    "Perishable" : "PER",
    "Print" : "PTR",
    "Processor" : "PROC",
    "Project based" : "PROB",
    "Screen" : "SCR",
    "Server" : "SER",
    "Time based" : "TIMB",
    "Transport" : "TRAN"
}
AXELOR_PRODUCT_FAMILIES = {
    "Components" : 3,
    "Consumables" : 4,
    "Equipment" : 2,
    "Expenses" : 6,
    "Services" : 5,
    "Subscription" : 1
}
AXELOR_PRODUCT_TYPE = [
    "Product",
    "Service"
]

parser = argparse.ArgumentParser(
    description="Converts inventory lists to a Axelor-compatible CSV format",
    epilog="Licensed under the 0BSD.")
parser.add_argument("-t", "--type", default="xslx",
    choices=["xslx"],
    help="The type of file that is being imported")
parser.add_argument("-v", "--version", action="version", version=ver_str)
parser.add_argument("-c", "--category", choices=AXELOR_PRODUCT_CATEGORIES, default="Package", help="Default product category to place in output")
parser.add_argument("-f", "--family", choices=AXELOR_PRODUCT_FAMILIES, default="Equipment", help="Default product family to place in output")
parser.add_argument("-T", "--Type", choices=AXELOR_PRODUCT_TYPE, default="Product", help="Default product type to place in output")
parser.add_argument("input", help="Input file")
parser_args = parser.parse_args()
input_file = parser_args.input
default_category = parser_args.category
default_family = parser_args.family
default_type = parser_args.Type

xslx_file = load_workbook(
    input_file,
    read_only=True,
    keep_vba=False,
    data_only=True,
    keep_links=False)

csv_out = csv.writer(sys.stdout, dialect="excel")
csv_out.writerow(AXELOR_CSV_COLUMNS)

# This is used to populate the CSV file.
xslx_col_index = {
    "description" : -1,
    "name" : -1,
    "category" : -1
}
xslx_headers = []

# A row with just a title would not fill up the entire
# max_column. As a result, there would be None at either
# the first or second position.
start_title_col = xslx_file.active.min_column
end_title_col = start_title_col + 1
# Assume the first line is not title unless otherwise found out.
xslx_header_row = xslx_file.active.min_row

for row in xslx_file.active.iter_rows(
    min_row=xslx_file.active.min_row,
    max_row=xslx_file.active.max_row,
    min_col=start_title_col,
    max_col=end_title_col,
    values_only=True
):
    for cell in row:
        if cell is None:
            xslx_header_row = xslx_header_row + 1

if xslx_header_row >= xslx_file.active.max_row:
    print("FE: Can't find headers in " + input_file + ".")
    sys.exit(1)

for row in xslx_file.active.iter_rows(
    min_row=xslx_header_row,
    max_row=xslx_header_row,
    min_col=xslx_file.active.min_column,
    max_col=xslx_file.active.max_column,
    values_only=True
):
    for cell in row:
        xslx_headers += [cell]
# col and row are indexed starting at 1.
header_index = 1
for header in xslx_headers:
    if (header.lower().find("descript") != -1) and (xslx_col_index["description"] == -1):
        xslx_col_index["description"] = header_index
    # The name would be one of the first columns in a file.
    # In case a file has a header containing "name" and another
    # containing "ID", it should still have the correct one.
    elif ((header.lower().find("name") != -1) or
    (header.lower().find("id") != -1)) and (xslx_col_index["name"] == -1):
        xslx_col_index["name"] = header_index
    elif (header.lower().find("cat") != -1) and (xslx_col_index["category"] == -1):
        xslx_col_index["category"] = header_index
    else:
        print("Warning: column " + '"' + header + '"' " from " + input_file +
        " will be ignored.", file=sys.stderr)
    header_index = header_index + 1
for required_header in xslx_col_index:
    if xslx_col_index[required_header] == -1:
        print("FE: " + required_header + " is not set", file=sys.stderr)
        sys.exit(1)

xslx_start_row = xslx_header_row + 1
xslx_min_col = -1
xslx_max_col = -1
for xslx_index in xslx_col_index.values():
    if xslx_index > xslx_max_col:
        xslx_max_col = xslx_index
    if (xslx_min_col == -1) or (xslx_index < xslx_min_col):
        xslx_min_col = xslx_index

cat_incr = {}
for row in xslx_file.active.iter_rows(
    min_row=xslx_start_row,
    max_row=xslx_file.active.max_row,
    min_col=xslx_min_col,
    max_col=xslx_max_col,
    values_only=True
):
    row_dict = {}
    for csv_col in AXELOR_CSV_COLUMNS:
        row_dict[csv_col] = ''
    col_index = xslx_min_col
    for cell in row:
        if col_index == xslx_col_index["name"]:
            #print("DEBUG: NAME") DEBUG
            row_dict["name"] = cell
            #print(row_dict["name"]) DEBUG
        elif col_index == xslx_col_index["description"]:
            row_dict["description"] = cell
            row_dict["internalDescription"] = cell
        elif col_index == xslx_col_index["category"]:
            valid_cat = False
            valid_fam = False
            for prod_cat in AXELOR_PRODUCT_CATEGORIES:
                if prod_cat == cell:
                    valid_cat = True
                    row_dict["productCategory_importId"] = AXELOR_PRODUCT_CATEGORIES[cell]
                    break
            for prod_fam in AXELOR_PRODUCT_FAMILIES:
                if prod_fam == cell:
                    valid_fam = True
                    row_dict["productFamily_importId"] = AXELOR_PRODUCT_FAMILIES[cell]
                    break
            code_incr = ''
            ax_code_str = ''
            if valid_fam:
                ax_code_str = AXELOR_FAMILY_SHORTHAND.get(cell, cell.upper().replace(" ", "_"))
            elif valid_cat:
                ax_code_str = AXELOR_CATEGORY_SHORTHAND.get(cell, cell.upper().replace(" ", "_"))
            else:
                ax_code_str = cell.upper().replace(" ", "_")
            code_incr = cat_incr.get(ax_code_str, "0000")
            cat_incr[ax_code_str] = code_incr
            row_dict["code"] = ax_code_str + "-" + code_incr
            if not valid_fam:
                row_dict["productFamily_importId"] = AXELOR_PRODUCT_FAMILIES[default_family]
            if not valid_cat:
                row_dict["productCategory_importId"] = AXELOR_PRODUCT_CATEGORIES[default_category]
            # Increment counter.
            code_incr_int = int(code_incr)
            code_incr_int = code_incr_int + 1
            if code_incr_int > 10:
                cat_incr[ax_code_str] = "".join('0' + '0' + '0' + str(code_incr_int))
            elif code_incr_int > 100:
                cat_incr[ax_code_str] = "".join('0' + '0' + str(code_incr_int))
            elif code_incr_int > 1000:
                cat_incr[ax_code_str] = "".join('0' + str(code_incr_int))
            else:
                cat_incr[ax_code_str] = str(code_incr_int)
        else:
            pass
        col_index = col_index + 1
    # Add values not directly based on XSLX file.
    row_dict["productTypeSelect"] = default_type
    row_dict["fullName"] = '[' + row_dict["code"] + ']' + " " + row_dict["name"]

    csv_out.writerow(row_dict.values())
    row_dict.clear()

xslx_file.close()