# Copyright 2021 Richard Johnston <techpowerawaits@outlook.com>
# SPDX-license-identifier: 0BSD

import string

from loguru import logger

try:
    import cell_pos
    from exceptions import InvconvMissingHeaders
    import ftype
    import msg_handler
except ModuleNotFoundError:
    import invconv.cell_pos as cell_pos
    from invconv.exceptions import InvconvMissingHeaders
    import invconv.ftype as ftype
    import invconv.msg_handler as msg_handler

used = True

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    used = False

# load_workbook is used repeatedly with similar settings
# every time.
WB_SETTINGS = {
    "read_only": True,
    "keep_vba": False,
    "data_only": True,
    "keep_links": False,
}


class XlsxDataTuple(ftype.BasicFtypeDataClass):
    def __init__(self, filename, wsname, headers):
        self.filename = filename
        self.wsname = wsname
        self.headers = headers
        self.cur_row = None
        self.cur_col = None
        super().__init__(
            filename=self.filename, sectionname=self.wsname, headers=self.headers
        )

    # Set relevant values and gets the number of operations
    # to be performed based on the dimensions.
    def set_oper_num(self, min_row, max_row, max_col):
        self.min_row = min_row
        self.min_col = 1
        self.max_row = max_row
        self.max_col = max_col

        delta_col = self.max_col - self.min_col + 1
        delta_row = self.max_row - self.min_row + 1
        self.num_oper = delta_col * delta_row
        return self.num_oper

    def load_workbook(self):
        return load_workbook(self.filename, **WB_SETTINGS)

    def parser(self):
        if self.cur_row is None:
            self.cur_row = self.min_row
        if self.cur_col is None:
            self.cur_col = self.min_col
        if self.cur_col > self.max_col:
            self.cur_col = self.min_col
            self.cur_row += 1
        if self.cur_row > self.max_row:
            self.cur_row = None
            self.cur_col = None
            return None
        col_letter = cell_pos.get_col_letter(self.cur_col)
        row_str = str(self.cur_row)
        wb = self.load_workbook()
        ws = wb[self.wsname]
        cell_val = ws[col_letter + row_str].value
        return_str = str(cell_val)
        if cell_val is None:
            return_str = ""
        if return_str == "#REF!":
            logger.warning(
                string.Template(
                    'Unknown reference found at $cell_pos in $id. Defaulting to "unknown".'
                ).substitute(
                    cell_pos=col_letter + row_str,
                    id=msg_handler.get_id((self.filename, self.wsname), "WS"),
                )
            )
            return_str = "unknown"
        self.cur_col += 1
        wb.close()
        return return_str


# Will store a file, worksheet tuple-like class
# with additional data accessible.
xlsx_data_list = ftype.FtypeDataList()
# Contains just a list of file, worksheet tuples.
xlsx_tuple_list = []

# xlsx files always start counting at 1.
INVALID_ROW = 0


def start(input_files):
    # Gets the name of worksheets and
    # adds it to xlsx_tuple_list.
    get_worksheets(input_files)
    # Sometimes, openpyxl can't get
    # the proper dimensions of a worksheet,
    # so it handles that. It also deals with
    # headers in the worksheets and removes
    # blank cells from the size of the sheet.
    set_data()
    # Check if some file worksheet pairs don't
    # have a valid header.
    if not xlsx_data_list:
        raise InvconvMissingHeaders
    # Can't directly check for membership of
    # items from xlsx_tuple_list in xlsx_data_list,
    # for they are different types.
    for file_section in xlsx_tuple_list:
        found_file_section = False
        for data_file_section in xlsx_data_list:
            # The first element in if statement
            # has to be XlsxDataTuple, as it
            # contains a __eq__() function
            # that should work in this case.
            if data_file_section == file_section:
                found_file_section = True
                break
        if not found_file_section:
            logger.error(
                f"{msg_handler.get_id(file_section, 'ws')} contains no valid headers."
            )
            msg_handler.does_continue()
    return xlsx_data_list


def get_worksheets(input_files):
    for input_file in input_files:
        wb = load_workbook(input_file, **WB_SETTINGS)
        sheetname_list = wb.sheetnames
        for sheetname in sheetname_list:
            xlsx_tuple_list.append((input_file, sheetname))
        wb.close()


def set_data():
    for filename, wsname in xlsx_tuple_list:
        wb = load_workbook(filename, **WB_SETTINGS)
        ws = wb[wsname]
        # max_col and max_row can be None.
        cur_max_col = ws.max_column
        cur_max_row = ws.max_row
        # Close workbook right away so
        # it won't remain open in case script
        # gets closed or crashes.
        wb.close()
        max_col = get_max_col(filename, wsname, cur_max_col)
        max_row = get_max_row(filename, wsname, cur_max_row)
        # Get the row where a header was found.
        header_row = get_header_row(filename, wsname, max_row)
        # check_header_row() ensures that a non-blank row
        # is after header row. If not, it might not
        # actually be a header row.
        if (
            header_row == INVALID_ROW
            or header_row == max_row
            or not check_header_row(filename, wsname, max_col, header_row)
        ):
            continue
        # The first row after the header_row.
        min_row = header_row + 1
        header_list = get_header_list(filename, wsname, max_col, header_row)
        if max_col > len(header_list):
            logger.info(
                string.Template(
                    "Reducing max column length of $id from $cur_col to $new_col due to None in $cell_pos."
                )
            )
            max_col = len(header_list)
        DataTuple = XlsxDataTuple(filename, wsname, header_list)
        DataTuple.set_oper_num(min_row, max_row, max_col)
        xlsx_data_list.append(DataTuple)


def get_max_col(filename, wsname, max_col):
    xlsx_id = msg_handler.get_id((filename, wsname), "WS")
    while (not isinstance(max_col, int)) or (max_col <= INVALID_ROW):
        logger.error(f"Max col for {xlsx_id} is {str(max_col)}.")
        msg_handler.does_continue()
        try:
            logger.info("User providing number of columns (starting at 1).")
            max_col = int(
                input("Please provide the number of columns (starting at 1) > ")
            )
        except (ValueError, TypeError):
            logger.log("FAILURE", "Input could not be converted to int.")
            max_col = None
        if (isinstance(max_col, int)) and (max_col <= 0):
            logger.log("FAILURE", "Input is less than one.")
    return max_col


def get_max_row(filename, wsname, max_row):
    xlsx_id = msg_handler.get_id((filename, wsname))
    while (not isinstance(max_row, int)) or (max_row <= 0):
        logger.error(f"Max row for {xlsx_id} is {str(max_row)}.")
        msg_handler.does_continue()
        try:
            logger.info("User providing number of rows (starting at 1).")
            max_row = int(input("Please provide the number of rows (starting at 1) > "))
        except (ValueError, TypeError):
            logger.log("FAILURE", "Input could not be converted to int.")
            max_row = None
        if (isinstance(max_row, int)) and (max_row <= 0):
            logger.log("FAILURE", "Input is less than one.")
    return max_row


def get_header_row(filename, wsname, max_row):
    wb = load_workbook(filename, **WB_SETTINGS)
    ws = wb[wsname]
    # header_row starts at 1,
    # so a value of 0 indicates
    # it wasn't found.
    header_row = INVALID_ROW

    for row in cell_pos.row_iter(max_row):
        row_str = str(row)
        # A row with just a title would not fill up the entire max_column.
        # As a result, there would be None at either the first or second
        # position.
        cell1 = ws["A" + row_str].value
        cell2 = ws["B" + row_str].value
        if cell1 is not None and cell2 is not None:
            header_row = row
            break
    wb.close()
    return header_row


def check_header_row(filename, wsname, max_col, header_row):
    wb = load_workbook(filename, **WB_SETTINGS)
    ws = wb[wsname]

    # Check the row after the header row
    # for content.
    post_header_row = header_row + 1
    row_str = str(post_header_row)
    # List of items in row.
    row_list = []

    for col in cell_pos.col_iter(max_col):
        col_letter = cell_pos.get_col_letter(col)
        row_list.append(str(ws[col_letter + row_str].value))
    wb.close()
    # Ensure the row is not blank.
    if row_list.count("None") != len(row_list):
        return True
    return False


def get_header_list(filename, wsname, max_col, header_row):
    wb = load_workbook(filename, **WB_SETTINGS)
    ws = wb[wsname]
    header_list = []

    row_str = str(header_row)
    for col in cell_pos.col_iter(max_col):
        col_letter = cell_pos.get_col_letter(col)
        header_item = ws[col_letter + row_str].value
        # Assuming the header doesn't have blank
        # items between entries. Only at the end.
        if header_item is None:
            logger.warning(
                f"Blank header {col_letter+row_str} in {msg_handler.get_id((filename, wsname), 'WS')} will be ignored."
            )
            break
        header_list.append(header_item)

    wb.close()
    return header_list


if used:
    ftype.add("xlsx", start)
