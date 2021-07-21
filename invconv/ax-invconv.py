#!/usr/bin/env python3

# Copyright 2021 Richard Johnston <techpowerawaits@outlook.com>
# SPDX-license-identifier: 0BSD

import argparse
import os.path
import string

from loguru import logger
from openpyxl import load_workbook
import progress.bar

import axm.parser
import axm.utils
import cell_pos
import common
import logic
import msg_handler

ver_path = os.path.join(os.path.pardir, "VERSION")
try:
    with open(ver_path, "r") as version_file:
        __version__ = version_file.readline()
except FileNotFoundError:
    __version__ = "Unknown"

parser = argparse.ArgumentParser(
    description="Converts inventory lists to a Axelor-compatible CSV format",
    epilog="Licensed under the 0BSD.",
    add_help=False,
)
parser.add_argument("-d", "--data-file", default="demo.ini", help="INI data file")
data_file = parser.parse_known_args()[0].data_file
with open(data_file) as data_fptr:
    common.init(data_fptr)

parser.add_argument("-m", "--map-file", default="default.axm", help="AXM map file")
parser.add_argument(
    "-l",
    "--log-file",
    default=msg_handler.get_default_logname(),
    help="File to store messages",
)
parser.add_argument("-D", "--debug", action="store_true", help="Enables debugging")
parser.add_argument(
    "-h", "--help", action="help", help="show this help message and exit"
)
parser.add_argument("-v", "--version", action="version", version=__version__)
parser.add_argument(
    "-t",
    "--type",
    default="xlsx",
    choices=["xlsx"],
    help="The type of file that is being imported",
)

# Looks at the fallback-related arguments defined in data file
# and officially adds it as an argument.
for section_name, arg_tuple in common.arg_dict.items():
    short_arg = arg_tuple.short
    long_arg = arg_tuple.long
    help_text = arg_tuple.help
    parser.add_argument(
        short_arg,
        long_arg,
        default=common.fallback[section_name],
        choices=common.meta_table[section_name],
        help=help_text,
    )

parser.add_argument("input", nargs="+", help="Input file(s)")
parser_dict = vars(parser.parse_args())
input_files = parser_dict["input"]
map_file = parser_dict["map_file"]
common.is_debug = parser_dict["debug"]

# Set up logger.
# (If any errors occured before
# this point, it would be handled
# by loguru's default log handler.)
msg_handler.init()
msg_handler.set_log(parser_dict["log_file"])

# Looks at the fallback-related arguments again and
# replaces all appropriate values in the fallback dict
# with the user-provided values. If the user didn't override
# the default fallback values, then the fallback values
# in the dict will be replaced with the values it currently has.
for section_name, arg_tuple in common.arg_dict.items():
    long_arg = arg_tuple.long
    key = long_arg.removeprefix("--").replace("-", "_")
    common.fallback[section_name] = parser_dict[key]

# On some xlsx files, the max_row and max_col
# cannot be read.
max_rows = {}
max_cols = {}

# Store dictionary of header names for future
# reference.
file_ws_dict = {}

for input_file in input_files:
    file_ws_dict[input_file] = []
    xlsx_file = load_workbook(
        input_file, read_only=True, keep_vba=False, data_only=True, keep_links=False
    )
    xlsx_ws_list = xlsx_file.sheetnames
    for ws_name in xlsx_ws_list:
        xlsx_id = msg_handler.get_xlsx_id(input_file, ws_name)
        file_ws_dict[input_file].append(ws_name)
        ws = xlsx_file[ws_name]

        max_row = ws.max_row
        max_col = ws.max_column
        while (not isinstance(max_row, int)) or (max_row <= 0):
            logger.error(f"Max row for {xlsx_id} is {str(max_row)}.")
            msg_handler.does_continue()
            try:
                logger.info("User providing number of rows (starting at 1).")
                max_row = int(
                    input("Please provide the number of rows (starting at 1) > ")
                )
            except (ValueError, TypeError):
                logger.log("FAILURE", "Input could not be converted to int.")
                max_row = None
            if (isinstance(max_row, int)) and (max_row <= 0):
                logger.log("FAILURE", "Input is less than one.")
        max_rows[(input_file, ws_name)] = max_row
        while (not isinstance(max_col, int)) or (max_col <= 0):
            logger.error(f"Max row for {xlsx_id} is {str(max_row)}")
            msg_handler.does_continue()
            try:
                logger.info("User providing number of columns (starting at 1).")
                max_col = int(
                    input("Please provide the number of columns (starting at 1) > ")
                )
            except (ValueError, TypeError):
                logger.log("FAILURE", "Input could not be converted to int.")
                max_col = None
            if (isinstance(max_col, int)) and (max_col <= 0):
                logger.log("FAILURE", "Input is less than one.")
        max_cols[(input_file, ws_name)] = max_col
    xlsx_file.close()

# The xlsx files might have a title,
# which needs to be avoided.
min_header_rows = {}

# Find where headers are inside each worksheet.
for input_file in file_ws_dict:
    xlsx_file = load_workbook(
        input_file, read_only=True, keep_vba=False, data_only=True, keep_links=False
    )
    for ws_name in file_ws_dict[input_file]:
        ws = xlsx_file[ws_name]
        header_row = 0

        for row in cell_pos.row_iter(max_rows[(input_file, ws_name)]):
            row_str = str(row)
            # A row with just a title would not fill up the entire max_column.
            # As a result, there would be None at either the first or second
            # position.
            cell1 = ws["A" + row_str].value
            cell2 = ws["B" + row_str].value
            if cell1 is not None and cell2 is not None:
                header_row = row
                break

        # Only add to min_header_row if the header was found
        # and is not the only thing in object.
        if 0 < header_row < max_rows[(input_file, ws_name)]:
            # Check if there are only empty rows after header.
            post_header = header_row + 1
            post_header_list = []
            for col_incr in cell_pos.col_iter(max_cols[(input_file, ws_name)]):
                col_letter = cell_pos.get_col_letter(col_incr)
                row_str = str(post_header)
                post_header_list.append(str(ws[col_letter + row_str].value))
            if post_header_list.count("None") != len(post_header_list):
                min_header_rows[(input_file, ws_name)] = header_row
    xlsx_file.close()

# Check if script can be continued.
if len(min_header_rows) == 0:
    logger.critical("No file contained valid headers.")

# Temp file list so that keys won't be deleted
# in the dictionary being parsed.
file_list = file_ws_dict.keys()

for input_file in file_list:
    is_file_used = False
    for used_file_ws in min_header_rows:
        used_file = used_file_ws[0]
        if input_file == used_file:
            is_file_used = True
            break
    if not is_file_used:
        logger.error(
            string.Template("$file contains no valid headers.").substitute(
                file=input_file
            )
        )
        del file_ws_dict[input_file]

# Make sure file_list isn't accidently used.
del file_list

for input_file in file_ws_dict:
    for ws_name in file_ws_dict[input_file]:
        is_ws_used = False
        for used_file_ws in min_header_rows:
            used_file = used_file_ws[0]
            used_ws = used_file_ws[1]
            if input_file == used_file and ws_name == used_ws:
                is_ws_used = True
                break
        if not is_ws_used:
            logger.panic(
                f"{msg_handler.get_xlsx_id(input_file, ws_name)} contains no valid headers."
            )
            file_ws_dict[input_file].remove(ws_name)

# Record all headers for remaining files and worksheets.
xlsx_headers = {}
for input_file in file_ws_dict:
    xlsx_file = load_workbook(
        input_file, read_only=True, keep_vba=False, data_only=True, keep_links=False
    )
    for ws_name in file_ws_dict[input_file]:
        xlsx_headers[(input_file, ws_name)] = []
        ws = xlsx_file[ws_name]
        # These variables are used to keep
        # track of if a valid string or int
        # appears after a set of Nones in the
        # header.
        start_none = 0
        valid_after_none = False
        for col_incr in cell_pos.col_iter(max_cols[(input_file, ws_name)]):
            col_letter = cell_pos.get_col_letter(col_incr)
            row_str = str(min_header_rows[(input_file, ws_name)])
            cell = ws[col_letter + row_str].value

            if cell is None:
                if start_none == 0:
                    start_none = col_incr
                logger.warning(
                    f"Blank header {col_letter+row_str} in {msg_handler.get_xlsx_id(input_file, ws_name)} will be ignored."
                )
            else:
                if start_none != 0 and not valid_after_none:
                    valid_after_none = True
                xlsx_headers[(input_file, ws_name)].append(str(cell))
        if start_none > 0 and not valid_after_none:
            before_none = start_none - 1
            if before_none == 0:
                logger.error(
                    string.Template(
                        "Attempted to reduce max column length of $id from $col to 0 due to None in $cell_pos."
                    )
                ).substitute(
                    col=max_cols[(input_file, ws_name)],
                    id=msg_handler.get_xlsx_id(input_file, ws_name),
                    cell_pos=cell_pos.get_col_letter(start_none)
                    + str(min_header_rows[(input_file, ws_name)]),
                )
            else:
                logger.info(
                    string.Template(
                        "Reducing max column length of $id from $cur_col to $new_col due to None in $cell_pos."
                    )
                ).substitute(
                    id=msg_handler.get_xlsx_id(input_file, ws_name),
                    cur_col=str(max_cols[(input_file, ws_name)]),
                    new_col=str(before_none),
                    cell_pos=cell_pos.get_col_letter(start_none)
                    + str(min_header_rows[(input_file, ws_name)]),
                )
                max_cols[(input_file, ws_name)] = before_none
    xlsx_file.close()

# Figure out the proper mapping between Axelor CSV and xlsx.
axm.parser.init(xlsx_headers)
with open(map_file) as map_fptr:
    while not axm.utils.is_eof(map_fptr):
        axm.parser.parse(map_fptr)
axm.parser.finalize()

# Setup progress bar.
max_oper = 0
for input_file in file_ws_dict:
    for ws_name in file_ws_dict[input_file]:
        # xlsx file parsing will start one row after header.
        net_row = (
            max_rows[(input_file, ws_name)] - min_header_rows[(input_file, ws_name)]
        )
        max_cells = net_row * max_cols[(input_file, ws_name)]
        max_oper += max_cells

# Convert xlsx to Axelor-compatible CSV.
logic.commit_headers()
with progress.bar.IncrementalBar(
    message="Generating output",
    max=max_oper,
    suffix="%(index)d/%(max)d ETA: %(eta_td)s",
) as progress_bar:
    for input_file in file_ws_dict:
        xlsx_file = load_workbook(
            input_file, read_only=True, keep_vba=False, data_only=True, keep_links=False
        )
        for ws_name in file_ws_dict[input_file]:
            ws = xlsx_file[ws_name]
            logic.init(input_file, ws_name, xlsx_headers[(input_file, ws_name)])

            # Don't need to start at header row.
            starting_row = min_header_rows[(input_file, ws_name)] + 1
            for row in cell_pos.row_iter(starting_row, max_rows[(input_file, ws_name)]):
                row_str = str(row)
                for col in cell_pos.col_iter(max_cols[(input_file, ws_name)]):
                    col_letter = cell_pos.get_col_letter(col)
                    cell = ws[col_letter + row_str].value
                    if cell == "#REF!":
                        logger.warning(
                            string.Template(
                                'Unknown reference found at $cell_pos in $id. Defaulting to "unknown".'
                            ).substitute(
                                cell_pos=col_letter + row_str,
                                id=msg_handler.get_xlsx_id(input_file, ws_name),
                            )
                        )
                        cell = "unknown"
                    logic.main(cell)
                    progress_bar.next()
        xlsx_file.close()
